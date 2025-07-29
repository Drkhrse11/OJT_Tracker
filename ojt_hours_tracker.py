import tkinter as tk
from tkinter import messagebox, simpledialog, Toplevel, Label
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os

FILENAME = "ojt_hours.xlsx"
TEMP_FILE = "temp_time_in.txt"

def setup_excel():
    if not os.path.exists(FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "OJT Logs"
        ws.append(["Date", "Day", "Time IN", "Time Out", "Total Hours"])
        wb.save(FILENAME)

def normalize_time_input(time_str):
    try:
        if ":" not in time_str:
            return None
        parts = time_str.strip().split(":")
        if len(parts) != 2:
            return None
        hour = int(parts[0])
        minute = int(parts[1])
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
        else:
            return None
    except:
        return None

def log_to_excel(date, day, time_in, time_out, total_hours):
    wb = load_workbook(FILENAME)
    ws = wb.active
    ws.append([date, day, time_in, time_out, total_hours])
    wb.save(FILENAME)

def save_time_in_to_file(date_str, time_in_str):
    with open(TEMP_FILE, "w") as f:
        f.write(f"{date_str},{time_in_str}")

def load_time_in_from_file():
    if os.path.exists(TEMP_FILE):
        with open(TEMP_FILE, "r") as f:
            content = f.read().strip()
            if content:
                parts = content.split(",")
                if len(parts) == 2:
                    return parts[0], parts[1]
    return None, None

def delete_temp_file():
    if os.path.exists(TEMP_FILE):
        os.remove(TEMP_FILE)

def time_in_action():
    time_in_str = simpledialog.askstring("Time In", "Enter Time In (e.g., 5:30 or 05:30):")
    if time_in_str:
        normalized = normalize_time_input(time_in_str)
        if not normalized:
            messagebox.showerror("Error", "Invalid time format. Use HH:MM.")
            return
        date_str = datetime.now().strftime("%Y-%m-%d")
        save_time_in_to_file(date_str, normalized)
        time_in_label.config(text=f"Time In: {normalized}")
        messagebox.showinfo("Saved", "Time In saved. You can close the app now.")
    else:
        messagebox.showwarning("Missing", "No Time In entered.")

def time_out_action():
    saved_date, saved_time_in = load_time_in_from_file()
    if not saved_time_in:
        messagebox.showwarning("Missing", "No saved Time In found. Please input Time In first.")
        return

    time_out_str = simpledialog.askstring("Time Out", "Enter Time Out (e.g., 17:45 or 5:45):")
    if time_out_str:
        normalized = normalize_time_input(time_out_str)
        if not normalized:
            messagebox.showerror("Error", "Invalid Time Out format. Use HH:MM.")
            return
        try:
            t_in = datetime.strptime(saved_time_in, "%H:%M")
            t_out = datetime.strptime(normalized, "%H:%M")
            if t_out < t_in:
                t_out += timedelta(days=1)
            total_hours = round((t_out - t_in).seconds / 3600, 2)

            day = datetime.now().strftime("%A")
            if day not in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
                messagebox.showwarning("Weekend", f"It's {day}. Only weekdays allowed.")
                return

            log_to_excel(saved_date, day, saved_time_in, normalized, total_hours)
            time_out_label.config(text=f"Time Out: {normalized}")
            total_label.config(text=f"Total Hours: {total_hours} hrs")
            time_in_label.config(text="Time In: ---")
            delete_temp_file()
            messagebox.showinfo("Logged", "Time In/Out saved to Excel.")
        except:
            messagebox.showerror("Error", "Something went wrong while calculating time.")
    else:
        messagebox.showwarning("Missing", "No Time Out entered.")

def edit_logs():
    wb = load_workbook(FILENAME)
    ws = wb.active
    edit_date = simpledialog.askstring("Edit Entry", "Enter date to edit (YYYY-MM-DD): ")
    found = False

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == edit_date:
            found = True
            new_time_in = simpledialog.askstring("Edit", f"Old Time In: {row[2].value}\nNew Time In (HH:MM): ")
            new_time_out = simpledialog.askstring("Edit", f"Old Time Out: {row[3].value}\nNew Time Out (HH:MM): ")
            new_in = normalize_time_input(new_time_in)
            new_out = normalize_time_input(new_time_out)
            if not new_in or not new_out:
                messagebox.showerror("Error", "Invalid time format.")
                return
            row[2].value = new_in
            row[3].value = new_out
            try:
                t_in = datetime.strptime(new_in, "%H:%M")
                t_out = datetime.strptime(new_out, "%H:%M")
                if t_out < t_in:
                    t_out += timedelta(days=1)
                total_hours = round((t_out - t_in).seconds / 3600, 2)
                row[4].value = total_hours
            except:
                messagebox.showerror("Error", "Time format error.")
            break

    if found:
        wb.save(FILENAME)
        messagebox.showinfo("Success", "Log updated.")
    else:
        messagebox.showwarning("Not found", "No entry found for that date.")

def delete_log_popup():
    def perform_delete():
        date_to_delete = date_entry.get()
        if not date_to_delete:
            messagebox.showwarning("Input Needed", "Enter a Date")
            return
        wb = load_workbook(FILENAME)
        ws = wb.active
        found = False
        for i in range(2, ws.max_row + 1):
            if ws[f"A{i}"].value == date_to_delete:
                ws.delete_rows(i)
                found = True
                break
        if found:
            wb.save(FILENAME)
            messagebox.showinfo("Deleted", f"Deleted log for {date_to_delete}")
            date_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Not found", "No log found for that date.")

    popup = Toplevel(root)
    popup.title("Delete Logs")
    popup.geometry("300x150")
    tk.Label(popup, text="Enter date (YYYY-MM-DD):", font=("Arial", 10)).pack(pady=5)
    date_entry = tk.Entry(popup, width=25)
    date_entry.pack(pady=5)
    tk.Button(popup, text="Delete", command=perform_delete).pack(padx=10)

def view_log():
    wb = load_workbook(FILENAME)
    ws = wb.active
    logs = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2] and row[3]:
            logs += f"{row[0]} ({row[1]}): ({row[2]}) - ({row[3]}) | {row[4]} hrs\n"

    if logs == "":
        logs = "No logs yet."

    log_window = Toplevel(root)
    log_window.title("Log Viewer")
    log_window.geometry("400x300")
    Label(log_window, text="OJT Logs", font=("Arial", 14, "bold")).pack(pady=5)
    Label(log_window, text=logs, justify="left", anchor="nw", font=("Courier", 10),
          wraplength=380).pack(padx=10, pady=10, fill="both", expand=True)

# --- App Start ---
setup_excel()
saved_date, saved_time = load_time_in_from_file()

root = tk.Tk()
root.title("OJT Hours Tracker")

tk.Label(root, text="OJT Hours Tracker", font=("Arial", 16)).pack(pady=10)

time_in_label = tk.Label(root, text=f"Time In: {saved_time if saved_time else '---'}", font=("Arial", 12))
time_in_label.pack()

time_out_label = tk.Label(root, text="Time Out: ---", font=("Arial", 12))
time_out_label.pack()

total_label = tk.Label(root, text="Total Hours: ---", font=("Arial", 12))
total_label.pack(pady=5)

btn_frame = tk.Frame(root)
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="Time In", width=10, command=time_in_action).grid(row=0, column=0, padx=5)
tk.Button(btn_frame, text="Time Out", width=10, command=time_out_action).grid(row=0, column=1, padx=5)
tk.Button(btn_frame, text="Edit Entry", width=10, command=edit_logs).grid(row=0, column=2, padx=5)
tk.Button(btn_frame, text="Delete Entry", width=10, command=delete_log_popup).grid(row=0, column=3, padx=5)
tk.Button(btn_frame, text="View Logs", width=10, command=view_log).grid(row=0, column=4, padx=5)

root.mainloop()
